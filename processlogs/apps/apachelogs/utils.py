from openpyxl import Workbook


def report(queryset):
	wb = Workbook()
	ws = wb.active
	ws.title = "Логи"

	fields = queryset.first()._meta.fields

	# Заполняем строку с названиями столбцов
	for idx, field in enumerate(fields):
		ws.cell(row=1, column=idx+1, value=field.verbose_name)

	# Наполняем отчет данными
	row = 2
	for log in queryset.iterator():
		for idx, field in enumerate(fields):
			ws.cell(row=row, column=idx + 1, value=getattr(log, field.attname))

		row += 1

	wb.save(filename='Report.xlsx')

	return wb
